# -*- coding: utf-8 -*-
from testcase import *
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN, BORDER_MEDIUM
import openpyxl as px
from openpyxl.worksheet.datavalidation import DataValidation


class TestCasePresenter:

    @staticmethod
    def output_test_result_report(test_cases: [TestCase], output_path: str) -> None:
        # Excelファイル生成
        workbook = px.Workbook()
        sheet = workbook.active
        sheet.title = 'テストケース'

        # 集計用エリアの高さ(単位: セル)
        note_cells_height = 4

        # 集計用エリアのヘッダーの位置
        note_formula_index = 3

        # 集計用エリアの集計値の位置
        note_header_index = 2

        # コンテンツのヘッダーの高さ
        contents_header_height = 1

        # コンテンツのヘッダーの位置
        contents_header_index = note_cells_height + contents_header_height

        # コンテンツの高さ
        contents_height = len(test_cases)

        # コンテンツの開始位置
        contents_begin_index = note_cells_height + contents_header_height + 1

        # コンテンツの終了位置
        contents_end_index = contents_height + contents_header_height + note_cells_height

        # 囲み罫線
        surround_bolder = Border(
            left=Side(style=BORDER_THIN),
            right=Side(style=BORDER_THIN),
            top=Side(style=BORDER_THIN),
            bottom=Side(style=BORDER_THIN)
        )

        # ヘッダーの背景色
        header_bg_color = PatternFill(
            patternType='solid',
            fgColor='006000'
        )

        # 集計用エリアを設定
        note_header = ["総項目数", "OK", "NG", "NT", "NA", "保留", "QA"]
        note_formula = ["=COUNTIF(I{0}:I{1}, \"*\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"OK\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"NG\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"NT\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"NA\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"保留\")".format(contents_begin_index, contents_end_index),
                        "=COUNTIF(J{0}:J{1}, \"QA\")".format(contents_begin_index, contents_end_index)]
        for col, header, formula in zip('JKLMNOP', note_header, note_formula):
            header_address = col + '{0}'.format(note_header_index)
            sheet[header_address] = header
            formula_address = col + '{0}'.format(note_formula_index)
            sheet[formula_address] = formula
            sheet[header_address].alignment = Alignment(vertical='center', horizontal='center', wrap_text=False)
            sheet[formula_address].alignment = Alignment(vertical='center', horizontal='center', wrap_text=False)
            sheet[header_address].font = Font(name='Yu Gothic', b=True, color='ffffff')
            sheet[header_address].fill = header_bg_color
            sheet[header_address].border = surround_bolder
            sheet[formula_address].border = surround_bolder

        # コンテンツのヘッダーを設定
        header = ["テスト対象", "テスト目的", "テスト条件", "前提条件", "手順", "期待値", "導出方法", "テストケース備考", "No", "ステータス", "端末", "ipaバージョン",
                  "実施者", "実施日", "バグID", "テスト結果備考"]
        for col, header in zip('ABCDEFGHIJKLMNOP', header):
            col_address = col + '{0}'.format(contents_header_index)
            sheet[col_address] = header
            sheet[col_address].alignment = Alignment(vertical='center', horizontal='center', wrap_text=False)
            sheet[col_address].font = Font(name='Yu Gothic', b=True, color='ffffff')
            sheet[col_address].fill = header_bg_color
            sheet[col_address].border = surround_bolder

        # テストケースを流し込む
        for index, test_case in enumerate(test_cases):
            col_address = index + contents_begin_index
            sheet["A{0}".format(col_address)] = test_case.test_target.value
            sheet["B{0}".format(col_address)] = test_case.test_purpose.value
            sheet["C{0}".format(col_address)] = test_case.test_condition.value
            sheet["D{0}".format(col_address)] = '\n'.join(map(lambda x: "・{0}".format(x.value), test_case.precondition))
            sheet["E{0}".format(col_address)] = '\n'.join(
                map(lambda x: "{0}. {1}".format(x[0] + 1, x[1].value), enumerate(test_case.procedure)))
            sheet["F{0}".format(col_address)] = '\n'.join(map(lambda x: "・{0}".format(x.value), test_case.expected_value))
            sheet["G{0}".format(col_address)] = '\n'.join(map(lambda x: "・{0}".format(x.value), test_case.derivation_method))
            sheet["H{0}".format(col_address)] = '\n'.join(map(lambda x: "・{0}".format(x.value), test_case.remarks))
            # Noを振る
            sheet["I{0}".format(col_address)] = index + 1

        # コンテンツの書式
        contents_range = sheet["A{0}:P{1}".format(contents_begin_index, contents_end_index)]
        for rows in contents_range:
            for cols in rows:
                cols.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
                cols.font = Font(name='Yu Gothic')
                cols.border = surround_bolder

        # コンテンツの「テスト対象」のマージ書式
        start = 2 + note_cells_height
        stop = 2 + note_cells_height
        for index, rows in enumerate(sheet["A{0}:A{1}".format(contents_begin_index, contents_end_index)]):
            target = sheet["A{0}".format(start)].value
            next = start + 1
            while target is not None and sheet["A{0}".format(next)].value == target:
                stop = next
                next += 1
            sheet.merge_cells("A{0}:A{1}".format(start, stop))
            start = next
            stop = next

        # FIXME: 大項目を跨いだ場合でも隣接する「テスト目的」が同じであればセル結合が実施されてしまう
        # コンテンツの「テスト目的」のマージ書式
        start = 2 + note_cells_height
        stop = 2 + note_cells_height
        for index, rows in enumerate(sheet["B{0}:B{1}".format(contents_begin_index, contents_end_index)]):
            target = sheet["B{0}".format(start)].value
            next = start + 1
            while target is not None and sheet["B{0}".format(next)].value == target:
                stop = next
                next += 1
            sheet.merge_cells("B{0}:B{1}".format(start, stop))
            start = next
            stop = next

        # テストケースとテスト結果報告との間に太い罫線で仕切りを入れる
        contents_range = sheet["H{0}:H{1}".format(contents_begin_index, contents_end_index)]
        for rows in contents_range:
            for cols in rows:
                cols.border = Border(
                    left=Side(style=BORDER_THIN),
                    right=Side(style=BORDER_MEDIUM),
                    top=Side(style=BORDER_THIN),
                    bottom=Side(style=BORDER_THIN)
                )

        # コンテンツの「ステータス」で入力できる項目を指定
        dv = DataValidation(type="list", formula1='"OK,NG,NA,NT,NA,保留,QA"')
        dv.ranges = "J{0}:J{1}".format(contents_begin_index, contents_end_index)
        sheet.add_data_validation(dv)

        # オートフィルタを設定
        sheet.auto_filter.ref = "A{0}:P{1}".format(contents_begin_index, contents_end_index)

        # シート全体のセル幅設定
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            # 文字数と実際の幅には誤差があるため，余裕を持たせる
            adjusted_width = (max_length + 2) * 1.5
            sheet.column_dimensions[col[0].column_letter].width = adjusted_width

        # 保存
        workbook.save(output_path)
        print("テスト結果報告書の生成が完了しました．")
